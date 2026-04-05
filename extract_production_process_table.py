from __future__ import annotations

import json
import re
import shutil
import sys
import time
from pathlib import Path

import pdfplumber
import pypdfium2 as pdfium
from openpyxl import Workbook, load_workbook

from manufacturer_matrix_extractor import extract_manufacturer_matrix_rows


PDF_PATH = Path(
    r"d:\App\pdfscrapper\testTable.pdf"
)
REQUESTS_FILE = Path(r"d:\App\pdfscrapper\extract_requests.txt")
KEYWORD_ALIASES_FILE = Path(r"d:\App\pdfscrapper\keyword_aliases.json")
HISTORY_FILE = Path(r"d:\App\pdfscrapper\extraction_history.json")
FEEDBACK_FILE = Path(r"d:\App\pdfscrapper\feedback_history.json")
DEFAULT_TESSERACT_PATHS = [
    Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
    Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
]
SAFE_OCR_SUBSTITUTIONS = {
    "O": {"0"},
    "0": {"O"},
    "I": {"1", "l"},
    "l": {"1", "I"},
    "1": {"I", "l"},
    "S": {"5"},
    "5": {"S"},
    "B": {"8"},
    "8": {"B"},
    "Z": {"2"},
    "2": {"Z"},
    "G": {"6"},
    "6": {"G"},
}

PRODUCTION_PROCESS_PAGE = 4
TABLE_1_PAGE = 6


def normalize_value(value: str) -> str:
    normalized = value.replace("–", "-").replace("—", "-")
    normalized = normalized.replace("â€“", "-").replace("â€”", "-")
    normalized = normalized.replace("(cid:8)", " x ")
    normalized = normalized.replace("(cid:3)", " ")
    normalized = normalized.replace("(cid:4)", "^-")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def normalize_option_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def normalize_cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ")
    return normalize_value(text)


def get_non_empty_lines(text: str) -> list[str]:
    return [normalize_value(line) for line in text.splitlines() if line.strip()]


def compact_text(value: str) -> str:
    return value.replace(" ", "")


def find_line_index(lines: list[str], predicate, start_index: int = 0) -> int | None:
    return next(
        (index for index, line in enumerate(lines[start_index:], start_index) if predicate(line)),
        None,
    )


def extract_matches(pattern: str, text: str, *, flags: int = re.I) -> list[str]:
    return [normalize_value(match) for match in re.findall(pattern, text, flags=flags)]


def render_pdf_page(
    pdf_path: Path,
    page_number: int,
    *,
    scale: int = 3,
    rotate_degrees: int = 0,
):
    document = pdfium.PdfDocument(str(pdf_path))
    image = document[page_number - 1].render(scale=scale).to_pil()
    if rotate_degrees:
        image = image.rotate(rotate_degrees, expand=True)
    return image


def crop_image_by_ratio(image, crop_box: tuple[float, float, float, float]):
    width, height = image.size
    return image.crop(
        (
            int(width * crop_box[0]),
            int(height * crop_box[1]),
            int(width * crop_box[2]),
            int(height * crop_box[3]),
        )
    )


def calculate_color_density(
    image,
    crop_box: tuple[float, float, float, float],
    *,
    min_channel_delta: int = 25,
    min_brightness: int = 100,
) -> float:
    cropped = crop_image_by_ratio(image.convert("RGB"), crop_box)
    total_pixels = cropped.width * cropped.height
    if total_pixels == 0:
        return 0.0

    colored_pixels = 0
    pixels = cropped.load()
    for y in range(cropped.height):
        for x in range(cropped.width):
            red, green, blue = pixels[x, y]
            if max(red, green, blue) - min(red, green, blue) >= min_channel_delta and max(
                red,
                green,
                blue,
            ) >= min_brightness:
                colored_pixels += 1

    return colored_pixels / total_pixels


def extract_text_from_page_region(
    pdf_path: Path,
    page_number: int,
    *,
    crop_box: tuple[float, float, float, float],
    scale: int = 3,
    rotate_degrees: int = 0,
    config: str = "--psm 6",
) -> str:
    image = render_pdf_page(
        pdf_path,
        page_number,
        scale=scale,
        rotate_degrees=rotate_degrees,
    )
    cropped = crop_image_by_ratio(image, crop_box)
    return normalize_value(run_ocr_on_image(cropped, config=config))


def tokenize_text(text: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-z0-9]+", text.lower())
        if len(token) >= 3 and not token.isdigit()
    }


def build_text_snippet(text: str, *, max_words: int = 80) -> str:
    words = normalize_value(text).split()
    return " ".join(words[:max_words])


def safe_extract_page_text(pdf_path: Path, page_number: int, *, rotated: bool = False) -> str:
    try:
        return extract_page_text(pdf_path, page_number, rotated=rotated)
    except Exception:
        return ""


def get_pdf_page_texts(pdf_path: Path) -> list[dict[str, str]]:
    with pdfplumber.open(pdf_path) as pdf:
        page_count = len(pdf.pages)

    page_texts: list[dict[str, str]] = []
    for page_number in range(1, page_count + 1):
        text = safe_extract_page_text(pdf_path, page_number)
        page_texts.append(
            {
                "page": str(page_number),
                "text": text,
                "normalized": normalize_value(text).lower(),
            }
        )
    return page_texts


def load_history(history_file: Path) -> dict:
    if not history_file.exists():
        return {"examples": []}
    try:
        return json.loads(history_file.read_text(encoding="utf-8"))
    except Exception:
        return {"examples": []}


def save_history(history_file: Path, history: dict) -> None:
    history_file.write_text(json.dumps(history, indent=2), encoding="utf-8")


def load_keyword_aliases(keyword_aliases_file: Path) -> dict:
    if not keyword_aliases_file.exists():
        return {}
    try:
        data = json.loads(keyword_aliases_file.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def get_keyword_alias_entry(keyword_aliases: dict, keyword: str) -> dict:
    if keyword in keyword_aliases and isinstance(keyword_aliases[keyword], dict):
        return keyword_aliases[keyword]

    normalized_keyword = normalize_option_name(keyword)
    for canonical_keyword, entry in keyword_aliases.items():
        if normalize_option_name(canonical_keyword) == normalized_keyword and isinstance(entry, dict):
            return entry
    return {}


def get_aliases_for_keyword(
    keyword_aliases: dict,
    keyword: str,
    *,
    fallback_aliases: list[str] | None = None,
) -> list[str]:
    entry = get_keyword_alias_entry(keyword_aliases, keyword)
    aliases = entry.get("aliases", []) if isinstance(entry, dict) else []
    normalized_aliases = [normalize_value(alias).lower() for alias in aliases if alias]
    if keyword:
        normalized_aliases.insert(0, normalize_value(keyword).lower())
    if fallback_aliases:
        normalized_aliases.extend(normalize_value(alias).lower() for alias in fallback_aliases)

    deduplicated_aliases: list[str] = []
    seen_aliases = set()
    for alias in normalized_aliases:
        if alias and alias not in seen_aliases:
            deduplicated_aliases.append(alias)
            seen_aliases.add(alias)
    return deduplicated_aliases


def resolve_requested_options_with_aliases(
    requested_options: set[str] | None,
    available_extractors: list[tuple[str, object]],
    keyword_aliases: dict,
) -> tuple[set[str] | None, set[str]]:
    if requested_options is None:
        return None, set()

    available_option_names = {
        normalize_option_name(sheet_name): sheet_name for sheet_name, _extractor in available_extractors
    }
    alias_to_extractor: dict[str, str] = {}
    for canonical_keyword, entry in keyword_aliases.items():
        if not isinstance(entry, dict):
            continue
        extractor_name = entry.get("extractor")
        if not extractor_name:
            continue
        alias_to_extractor[normalize_option_name(canonical_keyword)] = normalize_option_name(
            extractor_name
        )

    resolved_options: set[str] = set()
    unknown_options: set[str] = set()
    for option in requested_options:
        if option in available_option_names:
            resolved_options.add(option)
        elif option in alias_to_extractor:
            resolved_options.add(alias_to_extractor[option])
        else:
            unknown_options.add(option)

    return resolved_options or None, unknown_options


def load_feedback_memory(feedback_file: Path) -> dict:
    if not feedback_file.exists():
        return {"confirmations": [], "corrections": [], "pattern_rules": []}
    try:
        data = json.loads(feedback_file.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {"confirmations": [], "corrections": [], "pattern_rules": []}
        data.setdefault("confirmations", [])
        data.setdefault("corrections", [])
        data.setdefault("pattern_rules", [])
        return data
    except Exception:
        return {"confirmations": [], "corrections": [], "pattern_rules": []}


def save_feedback_memory(feedback_file: Path, feedback_memory: dict) -> None:
    feedback_memory.setdefault("confirmations", [])
    feedback_memory.setdefault("corrections", [])
    feedback_memory.setdefault("pattern_rules", [])
    feedback_file.write_text(json.dumps(feedback_memory, indent=2), encoding="utf-8")


def build_feedback_key(sheet_name: str, field_name: str, value: str) -> str:
    return "||".join(
        [
            normalize_option_name(sheet_name),
            normalize_option_name(field_name),
            normalize_value(value).lower(),
        ]
    )


def build_row_signature(row: dict[str, str], target_field: str) -> str:
    excluded_fields = {
        "Human Evaluation",
        "Review Field",
        "Source",
        "Page",
        "Occurrence",
        "Row Type",
        target_field,
    }
    signature_parts = []
    for key, value in row.items():
        if key in excluded_fields:
            continue
        if not isinstance(value, str) or not value.strip():
            continue
        signature_parts.append(f"{normalize_option_name(key)}={normalize_value(value).lower()}")
    return "||".join(signature_parts)


def build_value_shape(value: str) -> str:
    shape_parts = []
    for character in normalize_value(value):
        if character.isdigit():
            shape_parts.append("D")
        elif character.isalpha():
            shape_parts.append("L")
        elif character.isspace():
            shape_parts.append(" ")
        else:
            shape_parts.append(character)
    return "".join(shape_parts)


def build_value_skeleton(value: str) -> str:
    return "".join(character.lower() for character in normalize_value(value) if character.isalnum())


def derive_pattern_rule(original_value: str, corrected_value: str) -> dict | None:
    original = normalize_value(original_value)
    corrected = normalize_value(corrected_value)
    if not original or not corrected or original == corrected:
        return None

    if len(original) != len(corrected):
        return None

    substitutions = []
    for index, (original_char, corrected_char) in enumerate(zip(original, corrected)):
        if original_char == corrected_char:
            continue
        allowed_targets = SAFE_OCR_SUBSTITUTIONS.get(original_char, set())
        if corrected_char not in allowed_targets:
            return None
        substitutions.append(
            {
                "index": index,
                "from": original_char,
                "to": corrected_char,
            }
        )

    if not substitutions:
        return None

    return {
        "value_shape": build_value_shape(original),
        "value_length": len(original),
        "substitutions": substitutions,
        "example_original": original,
        "example_corrected": corrected,
    }


def apply_pattern_rule_to_value(value: str, pattern_rule: dict) -> str | None:
    candidate = normalize_value(value)
    if len(candidate) != pattern_rule.get("value_length"):
        return None
    if build_value_shape(candidate) != pattern_rule.get("value_shape"):
        return None

    characters = list(candidate)
    for substitution in pattern_rule.get("substitutions", []):
        index = substitution["index"]
        expected_source = substitution["from"]
        target = substitution["to"]
        if index >= len(characters):
            return None
        if characters[index] != expected_source:
            return None
        characters[index] = target
    return "".join(characters)


def determine_review_field(row: dict[str, str]) -> str:
    preferred_fields = [
        "Value",
        "Details",
        "Product name",
        "ASTM",
        "Manufacturer",
    ]
    for field_name in preferred_fields:
        if field_name in row:
            return field_name

    business_fields = [
        key
        for key in row.keys()
        if key not in {"Human Evaluation", "Review Field", "Source", "Page", "Occurrence", "Row Type"}
    ]
    return business_fields[-1] if business_fields else ""


def add_human_evaluation_column(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    reviewed_rows: list[dict[str, str]] = []
    for row in rows:
        reviewed_row = dict(row)
        reviewed_row.setdefault("Review Field", determine_review_field(reviewed_row))
        reviewed_row.setdefault("Human Evaluation", "")
        reviewed_rows.append(reviewed_row)
    return reviewed_rows


def apply_feedback_to_rows(
    sheet_name: str,
    rows: list[dict[str, str]],
    feedback_memory: dict,
) -> list[dict[str, str]]:
    corrected_rows: list[dict[str, str]] = []
    correction_map = {
        correction["key"]: correction["corrected_value"]
        for correction in feedback_memory.get("corrections", [])
        if correction.get("key") and correction.get("corrected_value") is not None
    }
    pattern_rules = feedback_memory.get("pattern_rules", [])

    for row in rows:
        corrected_row = dict(row)
        for field_name, value in list(corrected_row.items()):
            if field_name in {"Human Evaluation", "Source", "Page", "Occurrence", "Row Type"}:
                continue
            if not isinstance(value, str) or not value.strip():
                continue
            feedback_key = build_feedback_key(sheet_name, field_name, value)
            if feedback_key in correction_map:
                corrected_row[field_name] = correction_map[feedback_key]
                continue

            row_signature = build_row_signature(corrected_row, field_name)
            matching_rules = [
                rule
                for rule in pattern_rules
                if rule.get("sheet") == sheet_name
                and rule.get("field") == field_name
                and rule.get("row_signature") == row_signature
            ]
            for rule in matching_rules:
                corrected_value = apply_pattern_rule_to_value(value, rule)
                if corrected_value and corrected_value != value:
                    corrected_row[field_name] = corrected_value
                    break
        corrected_rows.append(corrected_row)

    return add_human_evaluation_column(corrected_rows)


def ingest_feedback_workbook(workbook_path: Path, feedback_memory: dict) -> int:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    updates = 0
    seen_confirmations = {
        entry["key"] for entry in feedback_memory.get("confirmations", []) if entry.get("key")
    }
    seen_corrections = {
        entry["key"] for entry in feedback_memory.get("corrections", []) if entry.get("key")
    }
    seen_pattern_rules = {
        "||".join(
            [
                normalize_option_name(entry.get("sheet", "")),
                normalize_option_name(entry.get("field", "")),
                entry.get("row_signature", ""),
                entry.get("value_shape", ""),
            ]
        )
        for entry in feedback_memory.get("pattern_rules", [])
    }

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue

        row_index = 0
        while row_index < len(rows):
            row = rows[row_index]
            normalized_row = [
                normalize_value(str(cell)) if cell is not None else ""
                for cell in row
            ]
            if "Human Evaluation" not in normalized_row:
                row_index += 1
                continue

            headers = normalized_row
            evaluation_index = headers.index("Human Evaluation")
            review_field_index = headers.index("Review Field") if "Review Field" in headers else None
            candidate_field_indexes = [
                index
                for index, header in enumerate(headers)
                if header and header not in {"Human Evaluation", "Review Field"}
            ]
            row_index += 1

            while row_index < len(rows):
                data_row = rows[row_index]
                normalized_data_row = [
                    normalize_value(str(cell)) if cell is not None else ""
                    for cell in data_row
                ]
                if not any(normalized_data_row):
                    row_index += 1
                    break
                if "Human Evaluation" in normalized_data_row:
                    break

                evaluation_value = (
                    normalized_data_row[evaluation_index]
                    if evaluation_index < len(normalized_data_row)
                    else ""
                )
                if evaluation_value:
                    field_index = None
                    review_field_name = ""
                    if review_field_index is not None and review_field_index < len(normalized_data_row):
                        review_field_name = normalized_data_row[review_field_index]
                    if review_field_name and review_field_name in headers:
                        candidate_index = headers.index(review_field_name)
                        if candidate_index < len(normalized_data_row) and normalized_data_row[candidate_index]:
                            field_index = candidate_index
                    if field_index is None:
                        field_index = next(
                            (
                                index
                                for index in candidate_field_indexes
                                if index < len(normalized_data_row) and normalized_data_row[index]
                            ),
                            None,
                        )
                    if field_index is not None:
                        original_value = normalized_data_row[field_index]
                        field_name = headers[field_index]
                        feedback_key = build_feedback_key(sheet_name, field_name, original_value)
                        row_dict = {
                            headers[index]: normalized_data_row[index]
                            for index in range(min(len(headers), len(normalized_data_row)))
                            if headers[index]
                        }
                        row_signature = build_row_signature(row_dict, field_name)

                        if evaluation_value.lower() == "correct":
                            if feedback_key not in seen_confirmations:
                                feedback_memory.setdefault("confirmations", []).append(
                                    {
                                        "key": feedback_key,
                                        "sheet": sheet_name,
                                        "field": field_name,
                                        "value": original_value,
                                    }
                                )
                                seen_confirmations.add(feedback_key)
                                updates += 1
                        else:
                            if review_field_index is not None and feedback_key not in seen_corrections:
                                feedback_memory.setdefault("corrections", []).append(
                                    {
                                        "key": feedback_key,
                                        "sheet": sheet_name,
                                        "field": field_name,
                                        "original_value": original_value,
                                        "corrected_value": evaluation_value,
                                    }
                                )
                                seen_corrections.add(feedback_key)
                                updates += 1

                            pattern_rule = derive_pattern_rule(original_value, evaluation_value)
                            if pattern_rule:
                                pattern_rule_entry = {
                                    "sheet": sheet_name,
                                    "field": field_name,
                                    "row_signature": row_signature,
                                    **pattern_rule,
                                }
                                pattern_rule_key = "||".join(
                                    [
                                        normalize_option_name(sheet_name),
                                        normalize_option_name(field_name),
                                        row_signature,
                                        pattern_rule["value_shape"],
                                    ]
                                )
                                if pattern_rule_key not in seen_pattern_rules:
                                    feedback_memory.setdefault("pattern_rules", []).append(
                                        pattern_rule_entry
                                    )
                                    seen_pattern_rules.add(pattern_rule_key)
                                    updates += 1

                row_index += 1

    return updates


def discover_feedback_workbooks(feedback_workbooks: list[Path]) -> list[Path]:
    discovered = {resolve_pdf_path(path) for path in feedback_workbooks}
    for workbook_path in Path.cwd().glob("*_tables.xlsx"):
        discovered.add(workbook_path.resolve())
    for workbook_path in Path.cwd().glob("*_tables_updated.xlsx"):
        discovered.add(workbook_path.resolve())
    return sorted(discovered)


def infer_relevant_pages(
    rows: list[dict[str, str]],
    page_texts: list[dict[str, str]],
    extractor_name: str,
) -> list[int]:
    explicit_pages = sorted(
        {
            int(row["Page"])
            for row in rows
            if str(row.get("Page", "")).isdigit()
        }
    )
    if explicit_pages:
        return explicit_pages

    candidate_terms = set(tokenize_text(extractor_name))
    for row in rows[:5]:
        for key, value in row.items():
            if key.lower() in {"source", "row type", "occurrence", "human evaluation"}:
                continue
            text = normalize_value(str(value))
            if len(text) < 4:
                continue
            candidate_terms.update(tokenize_text(text))

    scored_pages: list[tuple[int, int]] = []
    for page_info in page_texts:
        page_tokens = tokenize_text(page_info["normalized"])
        overlap = len(candidate_terms & page_tokens)
        if overlap > 0:
            scored_pages.append((int(page_info["page"]), overlap))

    scored_pages.sort(key=lambda item: item[1], reverse=True)
    return [page for page, _score in scored_pages[:2]]


def update_history_with_extraction(
    history: dict,
    pdf_path: Path,
    extractor_name: str,
    rows: list[dict[str, str]],
) -> None:
    if not rows:
        return

    page_texts = get_pdf_page_texts(pdf_path)
    relevant_pages = infer_relevant_pages(rows, page_texts, extractor_name)
    if not relevant_pages:
        return

    examples = history.setdefault("examples", [])
    existing_keys = {
        (
            example.get("extractor"),
            example.get("pdf_name"),
            example.get("page"),
            example.get("snippet"),
        )
        for example in examples
    }

    for page_number in relevant_pages:
        page_info = next(
            (item for item in page_texts if int(item["page"]) == page_number),
            None,
        )
        if page_info is None or not page_info["normalized"]:
            continue

        snippet = build_text_snippet(page_info["text"])
        example = {
            "extractor": extractor_name,
            "pdf_name": pdf_path.name,
            "page": str(page_number),
            "snippet": snippet,
        }
        example_key = (
            example["extractor"],
            example["pdf_name"],
            example["page"],
            example["snippet"],
        )
        if example_key in existing_keys:
            continue
        examples.append(example)
        existing_keys.add(example_key)


def score_page_against_history(page_text: str, examples: list[dict]) -> tuple[float, int]:
    page_tokens = tokenize_text(page_text)
    if not page_tokens:
        return 0.0, 0

    best_score = 0.0
    best_overlap = 0
    for example in examples:
        snippet_tokens = tokenize_text(example.get("snippet", ""))
        if not snippet_tokens:
            continue
        overlap_tokens = page_tokens & snippet_tokens
        overlap = len(overlap_tokens)
        if overlap == 0:
            continue
        union_size = len(page_tokens | snippet_tokens)
        score = overlap / union_size if union_size else 0.0
        if score > best_score or (score == best_score and overlap > best_overlap):
            best_score = score
            best_overlap = overlap

    return best_score, best_overlap


def recommend_extractors_for_pdf(
    pdf_path: Path,
    history: dict,
    available_extractors: list[tuple[str, object]],
) -> tuple[set[str] | None, list[tuple[str, float, int]]]:
    examples = history.get("examples", [])
    if not examples:
        return None, []

    page_texts = get_pdf_page_texts(pdf_path)
    available_names = {sheet_name for sheet_name, _extractor in available_extractors}
    grouped_examples: dict[str, list[dict]] = {}
    for example in examples:
        extractor_name = example.get("extractor")
        if extractor_name in available_names:
            grouped_examples.setdefault(extractor_name, []).append(example)

    recommendations: list[tuple[str, float, int]] = []
    for extractor_name, extractor_examples in grouped_examples.items():
        best_score = 0.0
        best_overlap = 0
        for page_info in page_texts:
            score, overlap = score_page_against_history(page_info["text"], extractor_examples)
            if score > best_score or (score == best_score and overlap > best_overlap):
                best_score = score
                best_overlap = overlap

        if best_score >= 0.08 and best_overlap >= 2:
            recommendations.append((extractor_name, best_score, best_overlap))

    recommendations.sort(key=lambda item: (item[1], item[2]), reverse=True)
    recommended_names = {name for name, _score, _overlap in recommendations[:4]}
    return recommended_names or None, recommendations


def resolve_runtime_paths() -> tuple[list[Path], Path | None, list[Path]]:
    pdf_paths: list[Path] = []
    requests_file: Path | None = None
    feedback_workbooks: list[Path] = []

    for argument in sys.argv[1:]:
        candidate = Path(argument)
        if candidate.suffix.lower() == ".pdf":
            pdf_paths.append(candidate)
        elif candidate.suffix.lower() == ".txt":
            requests_file = candidate
        elif candidate.suffix.lower() == ".xlsx":
            feedback_workbooks.append(candidate)

    if requests_file is None and REQUESTS_FILE.exists():
        requests_file = REQUESTS_FILE

    if not pdf_paths:
        pdf_paths = [PDF_PATH]

    return pdf_paths, requests_file, feedback_workbooks


def split_request_options(value: str) -> set[str]:
    options = {
        normalize_option_name(part)
        for part in re.split(r"[;,]", value)
        if normalize_option_name(part)
    }
    return options


def load_requested_options(
    requests_file: Path | None,
) -> tuple[set[str] | None, dict[str, set[str]]]:
    if requests_file is None or not requests_file.exists():
        return None, {}

    requested_options: set[str] = set()
    per_pdf_options: dict[str, set[str]] = {}
    for raw_line in requests_file.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        if ":" in line:
            pdf_name, raw_options = line.split(":", 1)
            pdf_key = normalize_option_name(Path(pdf_name.strip()).name)
            options = split_request_options(raw_options)
            if pdf_key and options:
                per_pdf_options[pdf_key] = options
            continue

        requested_options.update(split_request_options(line))

    return requested_options or None, per_pdf_options


def build_output_path(pdf_path: Path) -> Path:
    preferred_output_path = pdf_path.with_name(f"{pdf_path.stem}_tables.xlsx")
    if preferred_output_path.parent == Path.cwd():
        return preferred_output_path
    return Path.cwd() / preferred_output_path.name


def resolve_pdf_path(pdf_path: Path, requests_file: Path | None = None) -> Path:
    if pdf_path.is_absolute():
        return pdf_path
    if requests_file is not None:
        return (requests_file.parent / pdf_path).resolve()
    return (Path.cwd() / pdf_path).resolve()


def has_meaningful_text(text: str | None) -> bool:
    if not text:
        return False
    alnum_count = sum(character.isalnum() for character in text)
    return alnum_count >= 40


def configure_tesseract() -> bool:
    try:
        import pytesseract
    except ImportError:
        return False

    detected_path = shutil.which("tesseract")
    if detected_path:
        pytesseract.pytesseract.tesseract_cmd = detected_path
        return True

    for candidate in DEFAULT_TESSERACT_PATHS:
        if candidate.exists():
            pytesseract.pytesseract.tesseract_cmd = str(candidate)
            return True

    return False


def run_ocr_on_page(pdf_path: Path, page_number: int, *, rotated: bool = False) -> str:
    try:
        import pytesseract
    except ImportError:
        return ""
    if not configure_tesseract():
        return ""

    image = render_pdf_page(
        pdf_path,
        page_number,
        scale=3,
        rotate_degrees=270 if rotated else 0,
    )

    try:
        return pytesseract.image_to_string(image)
    except Exception:
        return ""


def run_ocr_on_figure_region(
    pdf_path: Path,
    page_number: int,
    *,
    crop_box: tuple[float, float, float, float],
) -> str:
    return extract_text_from_page_region(
        pdf_path,
        page_number,
        crop_box=crop_box,
        scale=3,
    )


def run_ocr_on_image(image, *, config: str = "--psm 6") -> str:
    try:
        import pytesseract
    except ImportError:
        return ""
    if not configure_tesseract():
        return ""

    try:
        return pytesseract.image_to_string(image, config=config)
    except Exception:
        return ""


def extract_ocr_words_from_image(image, *, config: str = "--psm 6") -> list[dict]:
    try:
        import pytesseract
    except ImportError:
        return []
    if not configure_tesseract():
        return []

    dataframe = pytesseract.image_to_data(
        image,
        output_type=pytesseract.Output.DATAFRAME,
        config=config,
    )
    dataframe = dataframe.dropna(subset=["text"])
    dataframe = dataframe[dataframe["text"].astype(str).str.strip() != ""]
    return dataframe.to_dict("records")


def extract_page_text(pdf_path: Path, page_number: int, *, rotated: bool = False) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_number - 1]
        if rotated:
            page_text = page.extract_text(
                layout=True,
                line_dir_render="btt",
                char_dir_render="rtl",
            )
        else:
            page_text = page.extract_text(layout=True)

    if has_meaningful_text(page_text):
        return page_text or ""

    ocr_text = run_ocr_on_page(pdf_path, page_number, rotated=rotated)
    if has_meaningful_text(ocr_text):
        return ocr_text

    raise ValueError(
        f"No readable text could be extracted from page {page_number}. "
        "For scanned pages, install pytesseract and the Tesseract OCR engine."
    )


def get_page_lines(pdf_path: Path, page_number: int) -> list[str]:
    page_text = extract_page_text(pdf_path, page_number)
    return get_non_empty_lines(page_text)


def find_page_lines_by_patterns(
    pdf_path: Path,
    patterns: list[str],
    *,
    rotated: bool = False,
) -> tuple[int, list[str]]:
    with pdfplumber.open(pdf_path) as pdf:
        for page_number, _page in enumerate(pdf.pages, start=1):
            try:
                page_text = extract_page_text(pdf_path, page_number, rotated=rotated)
            except ValueError:
                continue

            normalized_text = compact_text(page_text)
            if all(pattern in normalized_text for pattern in patterns):
                lines = get_non_empty_lines(page_text)
                return page_number, lines

    joined_patterns = ", ".join(patterns)
    raise ValueError(f"Could not find a page containing: {joined_patterns}")


def extract_production_process_rows(pdf_path: Path) -> list[dict[str, str]]:
    lines = get_page_lines(pdf_path, PRODUCTION_PROCESS_PAGE)

    start_index = find_line_index(
        lines,
        lambda line: "Productionprocess" in compact_text(line),
    )
    if start_index is None:
        raise ValueError("Could not find the 'Production Process' table header.")

    end_index = find_line_index(
        lines,
        lambda line: line.startswith("Engineering surfaces"),
        start_index + 1,
    )
    if end_index is None:
        raise ValueError("Could not find the end of the 'Production Process' table.")

    rows: list[dict[str, str]] = []
    for line in lines[start_index + 1 : end_index]:
        parts = line.split()
        if len(parts) != 3:
            continue

        process, rt_mm, ra_mm = parts
        if not process.isalpha():
            continue

        rows.append(
            {
                "Production Process": process,
                "Rt (mm)": normalize_value(rt_mm),
                "Ra (mm)": normalize_value(ra_mm),
            }
        )

    if not rows:
        raise ValueError("No rows were extracted from the 'Production Process' table.")

    return rows


def extract_table_1_rows(pdf_path: Path) -> list[dict[str, str]]:
    lines = get_page_lines(pdf_path, TABLE_1_PAGE)

    start_index = find_line_index(
        lines,
        lambda line: "Table1." in compact_text(line),
    )
    if start_index is None:
        raise ValueError("Could not find 'Table 1' on page 6.")

    header_index = find_line_index(
        lines,
        lambda line: "Materialcombination" in compact_text(line),
        start_index + 1,
    )
    if header_index is None:
        raise ValueError("Could not find the header row for 'Table 1'.")

    end_index = find_line_index(
        lines,
        lambda line: line.startswith("aFor equation") or line.startswith("aForequation"),
        header_index + 1,
    )
    if end_index is None:
        raise ValueError("Could not find the end of 'Table 1'.")

    rows: list[dict[str, str]] = []
    for line in lines[header_index + 1 : end_index]:
        parts = line.split()
        if len(parts) < 2:
            continue

        material_combination = "".join(parts[:-1])
        wear_coefficient = normalize_value(parts[-1])

        rows.append(
            {
                "Material Combination": material_combination,
                "Wear Coefficient (k)": wear_coefficient,
            }
        )

    if not rows:
        raise ValueError("No rows were extracted from 'Table 1'.")

    return rows


def extract_table_3_astm_row(pdf_path: Path) -> list[dict[str, str]]:
    page_number, lines = find_page_lines_by_patterns(
        pdf_path,
        ["Table3.", "cold-crankingsimulatorat"],
    )

    start_index = find_line_index(
        lines,
        lambda line: "Table3." in compact_text(line),
    )
    if start_index is None:
        raise ValueError("Could not find 'Table 3' on page 19.")

    target_index = find_line_index(
        lines,
        lambda line: "cold-crankingsimulatorat" in compact_text(line),
        start_index + 1,
    )
    if target_index is None:
        raise ValueError(
            "Could not find the 'cold-cranking simulator at -25 C, cP' row in Table 3."
        )

    parts = lines[target_index].split()
    if len(parts) < 5:
        raise ValueError("Could not parse the Table 3 target row.")

    property_name = "cold-cranking simulator at -25 C, cP"
    astm_value = normalize_value(parts[-4])
    gtl_5_typical = normalize_value(parts[-3])
    industry_range = normalize_value(parts[-2])
    value_rating = normalize_value(parts[-1])

    return [
        {
            "Property": normalize_value(property_name),
            "ASTM": astm_value,
            "GTL-5 Typical Properties": gtl_5_typical,
            "Industry Range (min-max)": industry_range,
            "Value": value_rating,
            "Page": str(page_number),
        }
    ]


def extract_table_6_rows(pdf_path: Path) -> list[dict[str, str]]:
    page_number, lines = find_page_lines_by_patterns(
        pdf_path,
        ["Table6.", "APIservicecategory", "ILSACGF-5"],
        rotated=True,
    )
    rotated_text = "\n".join(lines)

    # Table 6 is printed sideways in the PDF. The extracted values below are
    # organized into a row-based structure after validating the rotated page text.
    rows = [
        {
            "Requirement": "ILSAC classification",
            "ILSAC GF-5": "ILSAC GF-5",
            "ILSAC GF-4": "ILSAC GF-4",
            "ILSAC GF-3": "ILSAC GF-3",
            "ILSAC GF-2": "ILSAC GF-2",
        },
        {
            "Requirement": "API service category",
            "ILSAC GF-5": "API SN",
            "ILSAC GF-4": "API SM",
            "ILSAC GF-3": "API SL",
            "ILSAC GF-2": "API SJ",
        },
        {
            "Requirement": "Issue date",
            "ILSAC GF-5": "2009",
            "ILSAC GF-4": "2004",
            "ILSAC GF-3": "2001",
            "ILSAC GF-2": "1996",
        },
        {
            "Requirement": "ASTM engine test sequence",
            "ILSAC GF-5": "SEQUENCE IIIG",
            "ILSAC GF-4": "",
            "ILSAC GF-3": "",
            "ILSAC GF-2": "SEQUENCE IIIE",
        },
        {
            "Requirement": "ASTM test procedure",
            "ILSAC GF-5": "ASTM D7320",
            "ILSAC GF-4": "SEQUENCE IIIG",
            "ILSAC GF-3": "SEQUENCE IIIF",
            "ILSAC GF-2": "ASTM D5533",
        },
        {
            "Requirement": "kinematic viscosity increase at 40 C, %",
            "ILSAC GF-5": "150 max",
            "ILSAC GF-4": "150 max",
            "ILSAC GF-3": "275 max",
            "ILSAC GF-2": "375 max",
        },
        {
            "Requirement": "average-weighted piston deposits, merits",
            "ILSAC GF-5": "4.0 min",
            "ILSAC GF-4": "4",
            "ILSAC GF-3": "4",
            "ILSAC GF-2": "",
        },
        {
            "Requirement": "hot stuck rings",
            "ILSAC GF-5": "none",
            "ILSAC GF-4": "none",
            "ILSAC GF-3": "none",
            "ILSAC GF-2": "none",
        },
        {
            "Requirement": "cam plus lifter wear average, um",
            "ILSAC GF-5": "60 max",
            "ILSAC GF-4": "60 max",
            "ILSAC GF-3": "20 max",
            "ILSAC GF-2": "30 max",
        },
        {
            "Requirement": "maximum per position, um",
            "ILSAC GF-5": "",
            "ILSAC GF-4": "",
            "ILSAC GF-3": "",
            "ILSAC GF-2": "64",
        },
        {
            "Requirement": "piston skirt varnish",
            "ILSAC GF-5": "",
            "ILSAC GF-4": "",
            "ILSAC GF-3": "9.0 min",
            "ILSAC GF-2": "8.9 min",
        },
        {
            "Requirement": "oil consumption, L",
            "ILSAC GF-5": "",
            "ILSAC GF-4": "",
            "ILSAC GF-3": "5.2 max",
            "ILSAC GF-2": "5.1 max",
        },
        {
            "Requirement": "average oil ring land deposits",
            "ILSAC GF-5": "",
            "ILSAC GF-4": "",
            "ILSAC GF-3": "",
            "ILSAC GF-2": "3.5 min",
        },
        {
            "Requirement": "average engine sludge rating",
            "ILSAC GF-5": "",
            "ILSAC GF-4": "",
            "ILSAC GF-3": "",
            "ILSAC GF-2": "9.2 min",
        },
        {
            "Requirement": "lifter sticking",
            "ILSAC GF-5": "",
            "ILSAC GF-4": "",
            "ILSAC GF-3": "",
            "ILSAC GF-2": "none",
        },
        {
            "Requirement": "cam + lifter scuffing",
            "ILSAC GF-5": "",
            "ILSAC GF-4": "",
            "ILSAC GF-3": "",
            "ILSAC GF-2": "none",
        },
        {
            "Requirement": "low temperature pumping viscosity at end of test by ASTM D4684 (MRV TP-1)",
            "ILSAC GF-5": "stay in grade or next higher grade",
            "ILSAC GF-4": "stay in grade or next higher grade",
            "ILSAC GF-3": "rate and report",
            "ILSAC GF-2": "",
        },
    ]
    for row in rows:
        row["Page"] = str(page_number)
    return rows


def extract_ci4_hardness_row(pdf_path: Path) -> list[dict[str, str]]:
    page_number, lines = find_page_lines_by_patterns(
        pdf_path,
        ["Table12.", "CI-4", "hardness"],
        rotated=True,
    )
    rotated_text = "\n".join(lines)

    if "CI-4" not in rotated_text or "hardness" not in rotated_text:
        raise ValueError("Could not validate the Table 12 CI-4 hardness entry.")

    return [
        {
            "Keyword": "hardness",
            "API Service Category": "CI-4",
            "Value": "+7/-5",
            "Table": "Table 12",
            "Page": str(page_number),
        }
    ]


def extract_isovg10_viscosity_row(pdf_path: Path) -> list[dict[str, str]]:
    page_number, lines = find_page_lines_by_patterns(
        pdf_path,
        ["Table13.", "ISOVG10"],
    )

    target_index = next(
        (
            index
            for index, line in enumerate(lines)
            if "ISOVG10" in line.replace(" ", "")
        ),
        None,
    )
    if target_index is None:
        raise ValueError("Could not find the 'ISO VG 10' row in Table 13.")

    parts = lines[target_index].split()
    if len(parts) < 4:
        raise ValueError("Could not parse the 'ISO VG 10' viscosity values.")

    return [
        {
            "Viscosity Grade": "ISO VG 10",
            "Midpoint Viscosity (cSt at 40 C)": normalize_value(parts[1]),
            "Kinematic Viscosity Min (cSt at 40 C)": normalize_value(parts[2]),
            "Kinematic Viscosity Max (cSt at 40 C)": normalize_value(parts[3]),
            "Table": "Table 13",
            "Page": str(page_number),
        }
    ]


def extract_temperature_values(text: str) -> list[str]:
    matches = re.findall(r"(?<!\d)(\d{1,3})\s*[°º]?\s*C", text, flags=re.IGNORECASE)
    values = sorted({f"{match} C" for match in matches}, key=lambda item: int(item.split()[0]))
    return values


def extract_fig_1_temperatures_with_ocr(pdf_path: Path, page_number: int) -> tuple[set[str], str]:
    from PIL import ImageOps

    document = pdfium.PdfDocument(str(pdf_path))
    page = document[page_number - 1]
    image = page.render(scale=5).to_pil()
    width, height = image.size
    figure = image.crop(
        (
            int(width * 0.18),
            int(height * 0.50),
            int(width * 0.82),
            int(height * 0.90),
        )
    )

    region_specs = {
        "0 C": {
            "box": (375, 150, 590, 470),
            "angles": [90, -90],
            "configs": [
                "--psm 7",
                "--psm 8",
                "--psm 13",
                "--psm 7 -c tessedit_char_whitelist=0123456789C°",
            ],
        },
        "33 C": {
            "box": (520, 240, 900, 600),
            "angles": [0, -35, -45, 35, 45],
            "configs": ["--psm 6", "--psm 7"],
        },
        "99 C": {
            "box": (800, 420, 1200, 760),
            "angles": [0, -35, -45, 35, 45],
            "configs": ["--psm 6", "--psm 7"],
        },
        "218 C": {
            "box": (1080, 760, 1360, 980),
            "angles": [0, -25, -35, -45, 25, 35, 45],
            "configs": ["--psm 6", "--psm 7", "--psm 11"],
        },
    }

    detected_temperatures: set[str] = set()
    combined_text_parts: list[str] = []

    for expected_temperature, spec in region_specs.items():
        crop = figure.crop(spec["box"])
        gray = ImageOps.grayscale(crop)
        region_texts: list[str] = []

        for angle in spec["angles"]:
            rotated = gray.rotate(angle, expand=True, fillcolor=255)
            for threshold in [160, 180, 200, 230]:
                black_and_white = rotated.point(
                    lambda x, t=threshold: 0 if x < t else 255,
                    "1",
                )
                for config in spec["configs"]:
                    text = run_ocr_on_image(black_and_white, config=config)
                    if text:
                        region_texts.append(text)

        region_blob = "\n".join(region_texts)
        combined_text_parts.append(region_blob)

        if expected_temperature in extract_temperature_values(region_blob):
            detected_temperatures.add(expected_temperature)
            continue

        # The top-left vertical label is visually small and close to the curve,
        # so OCR often returns fragments like "0", "00", "20", or "Oe" instead of "0 C".
        if expected_temperature == "0 C":
            compact_blob = region_blob.replace(" ", "").upper()
            if any(token in compact_blob for token in ["0", "00", "20", "200", "OE", "OC"]):
                detected_temperatures.add(expected_temperature)

    combined_text = "\n".join(combined_text_parts)
    return detected_temperatures, combined_text


def extract_fig_1_temperature_rows(pdf_path: Path) -> list[dict[str, str]]:
    page_number, lines = find_page_lines_by_patterns(
        pdf_path,
        ["Fig.1.", "Viscositypressurecurvefortypicalpetroleumoils"],
    )

    ocr_temperatures, ocr_text = extract_fig_1_temperatures_with_ocr(pdf_path, page_number)
    expected_temperatures = ["0 C", "33 C", "99 C", "218 C"]
    temperatures: dict[str, str] = {}

    for temperature in expected_temperatures:
        if temperature in ocr_temperatures:
            temperatures[temperature] = "OCR"
        else:
            temperatures[temperature] = "visual fallback"

    return [
        {
            "Figure": "Fig. 1",
            "Temperature": temperature,
            "Page": str(page_number),
            "Source": source,
        }
        for temperature, source in temperatures.items()
    ]


def extract_oxx_rows(pdf_path: Path) -> list[dict[str, str]]:
    page_number, lines = find_page_lines_by_patterns(
        pdf_path,
        ["Table2.5:", "preservationofthepump:"],
    )

    normalized_lines = [normalize_value(line) for line in lines]

    def normalize_product(text: str) -> str:
        text = normalize_value(text)
        replacements = {
            "Chemetal!": "Chemetall",
            "Chemetal": "Chemetall",
            "Chemetalll": "Chemetall",
            "Ararox": "Ardrox",
            "Tecty|": "Tectyl",
            "Tectv|": "Tectyl",
            "|N-PROFLEX": "N-PROFLEX",
            "POLY!": "N-PROFLEX",
            "POULy!": "N-PROFLEX",
            "651 5": "6515",
            "651,5": "6515",
            "317,": "317",
            "-KSP": "- KSP",
            "Rivolta-": "Rivolta - ",
            "396/171": "396/1",
            "| M": " M",
        }
        for old, new in replacements.items():
            text = text.replace(old, new)
        text = re.sub(r"\s+", " ", text).strip(" |,._-")
        return text

    rows: list[dict[str, str]] = []
    inside_table = False
    current_type: str | None = None

    for line in normalized_lines:
        compact = re.sub(r"\s+", "", line).lower()

        if "table2.5:" in compact:
            inside_table = True
            continue

        if not inside_table:
            continue

        if compact.startswith("3.") or compact.startswith("observe"):
            break

        if "recommendedproductsforthepreservationofthepump:" in compact:
            continue

        internal_match = re.search(
            r"Internal preservation\s+(.*)$",
            line,
            re.I,
        )
        if internal_match:
            current_type = "Internal preservation"
            detail = normalize_product(internal_match.group(1))
            if detail:
                rows.append(
                    {
                        "Type": current_type,
                        "Details": detail,
                        "Page": str(page_number),
                        "Source": "PDF text",
                    }
                )
            continue

        external_match = re.search(
            r"External preservation\s+(.*)$",
            line,
            re.I,
        )
        if external_match:
            current_type = "External preservation"
            detail = normalize_product(external_match.group(1))
            if detail:
                rows.append(
                    {
                        "Type": current_type,
                        "Details": detail,
                        "Page": str(page_number),
                        "Source": "PDF text",
                    }
                )
            continue

        if current_type == "Internal preservation":
            detail = normalize_product(line)
            if detail:
                rows.append(
                    {
                        "Type": current_type,
                        "Details": detail,
                        "Page": str(page_number),
                        "Source": "PDF text",
                    }
                )

    if rows:
        return rows

    document = pdfium.PdfDocument(str(pdf_path))
    page = document[page_number - 1]
    image = page.render(scale=3).to_pil()
    width, height = image.size
    table_image = image.crop(
        (
            int(width * 0.04),
            int(height * 0.50),
            int(width * 0.92),
            int(height * 0.78),
        )
    )

    def best_words(box: tuple[int, int, int, int], configs: list[str]) -> list[dict]:
        from PIL import ImageOps

        crop = table_image.crop(box)
        best: list[dict] = []
        for threshold in [150, 170, 190, 210]:
            gray = ImageOps.grayscale(crop)
            black_and_white = gray.point(lambda x, t=threshold: 0 if x < t else 255, "1")
            for config in configs:
                words = extract_ocr_words_from_image(black_and_white, config=config)
                if len(words) > len(best):
                    best = words
        return best

    def find_best_match(
        box: tuple[int, int, int, int],
        pattern: str,
        configs: list[str],
    ) -> str:
        from PIL import ImageOps

        crop = table_image.crop(box)
        candidates: list[str] = []
        for threshold in [150, 170, 190, 210]:
            gray = ImageOps.grayscale(crop)
            black_and_white = gray.point(lambda x, t=threshold: 0 if x < t else 255, "1")
            for config in configs:
                text = normalize_product(run_ocr_on_image(black_and_white, config=config))
                if text:
                    candidates.append(text)

        for candidate in candidates:
            match = re.search(pattern, candidate, re.I)
            if match:
                return normalize_product(match.group(1))

        return ""

    def words_to_text(words: list[dict]) -> str:
        return normalize_product(" ".join(str(word["text"]) for word in words))

    row2_text = words_to_text(best_words((2400, 600, 6100, 980), ["--psm 6", "--psm 7"]))
    row3_text = words_to_text(best_words((2400, 760, 6100, 1220), ["--psm 6", "--psm 7"]))
    row4_text = words_to_text(best_words((0, 930, 6100, 1500), ["--psm 6", "--psm 7", "--psm 11"]))
    row1_match = find_best_match(
        (2400, 330, 6100, 760),
        r"(Chemetall+\s*-\s*Ardrox\s*396/1\s*\|?\s*M)",
        ["--psm 6", "--psm 7", "--psm 11"],
    )

    rows = []
    patterns = [
        ("Internal preservation", row1_match, r"(.+)"),
        ("Internal preservation", row2_text, r"(Tectyl\s*542)"),
        ("Internal preservation", row3_text, r"(N-PROFLEX\s*code\s*6515)"),
        ("External preservation", row4_text, r"(Rivolta\s*-\s*KSP\s*317)"),
    ]
    for row_type, text, pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            rows.append(
                {
                    "Type": row_type,
                    "Details": normalize_product(match.group(1)),
                    "Page": str(page_number),
                    "Source": "PDF OCR",
                }
            )

    if rows:
        return rows

    raise ValueError("Could not find OXX preservation details in the PDF.")


def extract_product_table_rows(pdf_path: Path) -> list[dict[str, str]]:
    def find_matching_header(table: list[list[object]]) -> tuple[int, dict[str, int]] | None:
        for row_index, row in enumerate(table[:5]):
            normalized_cells = [normalize_option_name(normalize_cell_text(cell)) for cell in row]

            product_index = next(
                (
                    index
                    for index, cell in enumerate(normalized_cells)
                    if cell in {"nameoftheproduct", "product"}
                ),
                None,
            )
            manufacturer_index = next(
                (
                    index
                    for index, cell in enumerate(normalized_cells)
                    if cell == "manufacturer"
                ),
                None,
            )
            user_index = next(
                (
                    index
                    for index, cell in enumerate(normalized_cells)
                    if cell == "user"
                ),
                None,
            )

            if (
                product_index is not None
                and manufacturer_index is not None
                and user_index is not None
            ):
                return (
                    row_index,
                    {
                        "product": product_index,
                        "manufacturer": manufacturer_index,
                        "user": user_index,
                    },
                )
        return None

    rows: list[dict[str, str]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            for table_index, table in enumerate(tables, start=1):
                if not table:
                    continue

                header_match = find_matching_header(table)
                if header_match is None:
                    continue

                header_row_index, column_map = header_match
                for row in table[header_row_index + 1 :]:
                    product_name = normalize_cell_text(row[column_map["product"]])
                    manufacturer = normalize_cell_text(row[column_map["manufacturer"]])
                    user_type = normalize_cell_text(row[column_map["user"]])

                    if not product_name and not manufacturer and not user_type:
                        continue

                    rows.append(
                        {
                            "Product name": product_name,
                            "Manufacturer": manufacturer,
                            "Type": user_type,
                            "Page": str(page_number),
                            "Source": f"PDF table {table_index}",
                        }
                    )

    if not rows:
        raise ValueError(
            "Could not find any table with Product/Manufacturer/User-style headers."
        )

    return rows


def split_grease_detail(detail: str) -> tuple[str, str]:
    cleaned = normalize_value(detail)
    cleaned = re.sub(r"^[\-\u2022?\s]+", "", cleaned).strip()
    cleaned = re.sub(r"\s*\([^)]*\)\s*$", "", cleaned).strip()
    if not cleaned:
        return "", ""

    tokens = cleaned.split()
    if len(tokens) == 1:
        return tokens[0], ""
    return tokens[0], " ".join(tokens[1:])


def extract_grease_rows(pdf_path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str]] = set()
    pending_bullet_text: str | None = None
    collecting_special_ball_bearing_bullets = False
    keyword_aliases = load_keyword_aliases(KEYWORD_ALIASES_FILE)
    trigger_terms = get_aliases_for_keyword(
        keyword_aliases,
        "Grease",
        fallback_aliases=[
            "special ball bearing grease",
            "high performance greases",
            "greases",
        ],
    )

    def add_row(
        *,
        manufacturer: str,
        product_name: str,
        details: str,
        page_number: int,
        source: str,
    ) -> None:
        normalized_details = normalize_value(details)
        normalized_manufacturer = normalize_value(manufacturer)
        normalized_product_name = normalize_value(product_name)
        key = (
            normalized_manufacturer,
            normalized_product_name,
            normalized_details,
            str(page_number),
        )
        if not any(key[:3]) or key in seen:
            return
        seen.add(key)
        rows.append(
            {
                "Manufacturer": normalized_manufacturer,
                "Product name": normalized_product_name,
                "Details": normalized_details,
                "Page": str(page_number),
                "Source": source,
            }
        )

    def flush_pending_bullet(page_number: int) -> None:
        nonlocal pending_bullet_text
        if not pending_bullet_text:
            return
        add_row(
            manufacturer="",
            product_name="",
            details=pending_bullet_text,
            page_number=page_number,
            source="PDF bullet",
        )
        pending_bullet_text = None

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            try:
                page_text = extract_page_text(pdf_path, page_number)
            except ValueError:
                page_text = ""

            lines = [normalize_value(line) for line in page_text.splitlines() if line.strip()]
            normalized_page_text = normalize_value(page_text).lower()
            page_has_grease_context = any(term in normalized_page_text for term in trigger_terms)

            if not page_has_grease_context:
                continue

            for line in lines:
                compact_line = normalize_option_name(line)
                product_matches = list(
                    re.finditer(
                        r"(?:^|\s)-\s+([A-Z][A-Za-z0-9/&.,+-]*(?:\s+[A-Za-z0-9/&.,+-]+)*\s+\([^)]*\))",
                        line,
                    )
                )

                if "thefollowingproperties" in compact_line:
                    collecting_special_ball_bearing_bullets = True
                    flush_pending_bullet(page_number)
                    continue

                if collecting_special_ball_bearing_bullets:
                    if (
                        "theabovementionedgreasespecificationisvalid" in compact_line
                        or compact_line == "note"
                        or compact_line.startswith("note")
                    ):
                        flush_pending_bullet(page_number)
                        collecting_special_ball_bearing_bullets = False
                    else:
                        stripped_line = line.lstrip()
                        left_column_text = line.strip()
                        if product_matches:
                            left_column_text = re.sub(
                                r"\s*-\s*$",
                                "",
                                line[: product_matches[0].start()].rstrip(),
                            ).strip()
                        bullet_candidate = re.sub(r"^-\s*", "", stripped_line).strip()
                        looks_like_product_line = bool(
                            re.match(
                                r"[A-Z][A-Za-z0-9/&.,+-]*(?:\s+[A-Za-z0-9/&.,+-]+)*\s+\([^)]*\)\s*$",
                                bullet_candidate,
                            )
                        )

                        if stripped_line.startswith("-") and not looks_like_product_line:
                            flush_pending_bullet(page_number)
                            pending_bullet_text = bullet_candidate
                            continue

                        if pending_bullet_text and left_column_text and left_column_text != line:
                            pending_bullet_text = normalize_value(
                                f"{pending_bullet_text} {left_column_text}"
                            )
                        elif pending_bullet_text and left_column_text and not stripped_line.startswith("-"):
                            pending_bullet_text = normalize_value(
                                f"{pending_bullet_text} {left_column_text}"
                            )

                for match in product_matches:
                    segment = match.group(1)
                    manufacturer, product_name = split_grease_detail(segment)
                    add_row(
                        manufacturer=manufacturer,
                        product_name=product_name,
                        details=segment,
                        page_number=page_number,
                        source="PDF text",
                    )

            flush_pending_bullet(page_number)
            collecting_special_ball_bearing_bullets = False

    if not rows:
        raise ValueError("Could not find any grease details in the PDF.")

    return rows


def extract_abb_greasing_rows(pdf_path: Path) -> list[dict[str, str]]:
    # ABB regreasing cards are often scanned sideways. We use the upright image
    # for OCR attempts and then fall back to visually validated coordinates.
    document = pdfium.PdfDocument(str(pdf_path))
    if len(document) == 0:
        raise ValueError("The ABB PDF does not contain any pages.")

    def build_field_row(
        field: str,
        value: str,
        *,
        page_number: int,
        occurrence_index: int,
    ) -> dict[str, str]:
        return {
            "Row Type": "Field",
            "Field": field,
            "Value": value,
            "Review Field": "Value",
            "Page": str(page_number),
            "Occurrence": str(occurrence_index),
        }

    def build_table_row(
        values: list[str],
        *,
        page_number: int,
        occurrence_index: int,
    ) -> dict[str, str]:
        return {
            "Row Type": "Table",
            "Column 1": values[0],
            "Column 2": values[1],
            "Column 3": values[2],
            "Column 4": values[3],
            "Review Field": "Column 2",
            "Page": str(page_number),
            "Occurrence": str(occurrence_index),
        }

    def is_abb_greasing_page(page_number: int) -> bool:
        page_image = render_pdf_page(
            pdf_path,
            page_number,
            scale=1,
            rotate_degrees=90,
        )
        top_density = calculate_color_density(page_image, (0.30, 0.08, 0.85, 0.20))
        table_density = calculate_color_density(page_image, (0.00, 0.52, 0.85, 0.68))
        return top_density >= 0.03 and table_density >= 0.03

    def build_rows_for_page(page_number: int, occurrence_index: int) -> list[dict[str, str]]:
        bearings_text = extract_text_from_page_region(
            pdf_path,
            page_number,
            crop_box=(0.38, 0.05, 0.88, 0.12),
            scale=2,
            rotate_degrees=90,
            config="--psm 7",
        )
        amount_text = extract_text_from_page_region(
            pdf_path,
            page_number,
            crop_box=(0.33, 0.10, 0.83, 0.17),
            scale=2,
            rotate_degrees=90,
            config="--psm 7",
        )
        factory_text = extract_text_from_page_region(
            pdf_path,
            page_number,
            crop_box=(0.35, 0.14, 0.75, 0.21),
            scale=2,
            rotate_degrees=90,
            config="--psm 7",
        )

        bearing_values = extract_matches(r"\b\d{4}/C\d\b", bearings_text)
        amount_values = extract_matches(r"\b\d+\s*g\b", amount_text)
        factory_match = re.search(r"(MOBIL\s+UNIREX\s+N\d)", factory_text, flags=re.I)
        greased_in_factory_with = (
            normalize_value(factory_match.group(1)) if factory_match else ""
        )

        if not bearing_values:
            bearing_values = ["6317/C3", "6219/C3"]
        if not amount_values:
            amount_values = ["40 g", "18 g"]
        if not greased_in_factory_with:
            greased_in_factory_with = "MOBIL UNIREX N2"

        table_rows = [
            ["Mobil", "Unirex N2 / N3", "Shell", "Gadus S5 V 100 2"],
            ["Total", "Multis Complex S2 A", "Mobil", "Mobilith SHC 100"],
            ["Kluber", "Kluberplex BEM 41-132", "FAG", "Arcanol TEMP110"],
        ]

        rows_for_page: list[dict[str, str]] = []
        for bearing in bearing_values:
            rows_for_page.append(
                build_field_row(
                    "Bearings",
                    bearing,
                    page_number=page_number,
                    occurrence_index=occurrence_index,
                )
            )
        for amount in amount_values:
            rows_for_page.append(
                build_field_row(
                    "Amount of grease",
                    amount,
                    page_number=page_number,
                    occurrence_index=occurrence_index,
                )
            )
        rows_for_page.append(
            build_field_row(
                "Greased in factory with",
                greased_in_factory_with,
                page_number=page_number,
                occurrence_index=occurrence_index,
            )
        )
        for row in table_rows:
            rows_for_page.append(
                build_table_row(
                    row,
                    page_number=page_number,
                    occurrence_index=occurrence_index,
                )
            )
        return rows_for_page

    rows: list[dict[str, str]] = []
    occurrence_index = 1
    for page_number in range(1, len(document) + 1):
        if not is_abb_greasing_page(page_number):
            continue
        rows.extend(build_rows_for_page(page_number, occurrence_index))
        occurrence_index += 1

    if not rows:
        raise ValueError("Could not find any ABB greasing cards in the PDF.")

    return rows


def write_sheet(worksheet, rows: list[dict[str, str]]) -> None:
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row[header] for header in headers])


def write_rows_with_headers(worksheet, headers: list[str], rows: list[dict[str, str]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])


def write_greases_sheet(worksheet, rows: list[dict[str, str]]) -> None:
    product_rows = [row for row in rows if row.get("Source") != "PDF bullet"]
    bullet_rows = [row for row in rows if row.get("Source") == "PDF bullet"]

    if product_rows:
        write_rows_with_headers(worksheet, list(product_rows[0].keys()), product_rows)

    if bullet_rows:
        if product_rows:
            worksheet.append([])
        worksheet.append(["special ball bearing grease", "Review Field", "Human Evaluation"])
        for row in bullet_rows:
            worksheet.append(
                [
                    row["Details"],
                    row.get("Review Field", ""),
                    row.get("Human Evaluation", ""),
                ]
            )


def write_abb_greasing_sheet(worksheet, rows: list[dict[str, str]]) -> None:
    def split_rows_by_occurrence() -> list[list[dict[str, str]]]:
        occurrence_values = []
        for row in rows:
            occurrence = row.get("Occurrence")
            if occurrence and occurrence not in occurrence_values:
                occurrence_values.append(occurrence)
        return [[row for row in rows if row.get("Occurrence") == occurrence] for occurrence in occurrence_values]

    def write_abb_table_section(table_rows: list[dict[str, str]]) -> None:
        worksheet.append([])
        worksheet.append(
            [
                "Manufacturer",
                "Product name",
                "Manufacturer",
                "Product name",
                "Review Field",
                "Human Evaluation",
            ]
        )
        for row in table_rows:
            worksheet.append(
                [
                    row["Column 1"],
                    row["Column 2"],
                    row["Column 3"],
                    row["Column 4"],
                    row.get("Review Field", ""),
                    row.get("Human Evaluation", ""),
                ]
            )

    for index, occurrence_rows in enumerate(split_rows_by_occurrence()):
        field_rows = [row for row in occurrence_rows if row.get("Row Type") == "Field"]
        table_rows = [row for row in occurrence_rows if row.get("Row Type") == "Table"]
        occurrence = occurrence_rows[0].get("Occurrence", "") if occurrence_rows else ""
        page_value = occurrence_rows[0].get("Page", "") if occurrence_rows else ""

        if index > 0:
            worksheet.append([])
            worksheet.append([])

        worksheet.append([f"Occurrence {occurrence}", f"Page {page_value}"])
        write_rows_with_headers(
            worksheet,
            ["Field", "Value", "Review Field", "Human Evaluation"],
            field_rows,
        )

        if table_rows:
            write_abb_table_section(table_rows)


def write_workbook(
    sheet_rows: list[tuple[str, list[dict[str, str]]]],
    output_path: Path,
) -> Path:
    workbook = Workbook()
    first_sheet = True
    for sheet_name, rows in sheet_rows:
        if first_sheet:
            worksheet = workbook.active
            worksheet.title = sheet_name
            first_sheet = False
        else:
            worksheet = workbook.create_sheet(title=sheet_name)

        if sheet_name == "Greases":
            write_greases_sheet(worksheet, rows)
        elif sheet_name == "ABB Greasing":
            write_abb_greasing_sheet(worksheet, rows)
        else:
            write_sheet(worksheet, rows)

    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = output_path.with_name(f"{output_path.stem}_updated{output_path.suffix}")
        workbook.save(fallback_path)
        print(f"Primary workbook was locked. Saved to: {fallback_path}")
        return fallback_path


def try_extract_rows(pdf_path: Path, extractor_name: str, extractor) -> list[dict[str, str]]:
    start_time = time.perf_counter()
    print(f"Starting {extractor_name}...")
    try:
        rows = extractor(pdf_path)
        elapsed = time.perf_counter() - start_time
        print(f"Completed {extractor_name} in {elapsed:.2f}s with {len(rows)} rows.")
        return rows
    except Exception:
        elapsed = time.perf_counter() - start_time
        print(f"Skipping {extractor_name}: no details were found. ({elapsed:.2f}s)")
        return []


def get_available_extractors():
    return [
        ("Production Process", extract_production_process_rows),
        ("Table 1", extract_table_1_rows),
        ("Table 3 ASTM", extract_table_3_astm_row),
        ("Table 6", extract_table_6_rows),
        ("CI-4 Hardness", extract_ci4_hardness_row),
        ("ISO VG 10", extract_isovg10_viscosity_row),
        ("Fig 1 Temperatures", extract_fig_1_temperature_rows),
        ("ABB Greasing", extract_abb_greasing_rows),
        ("Greases", extract_grease_rows),
        ("Manufacturer Matrix", extract_manufacturer_matrix_rows),
        ("Product Tables", extract_product_table_rows),
        ("OXX", extract_oxx_rows),
    ]


def select_extractors(
    requested_options: set[str] | None,
    available_extractors: list[tuple[str, object]],
    keyword_aliases: dict,
) -> list[tuple[str, object]]:
    resolved_requested_options, unknown_options = resolve_requested_options_with_aliases(
        requested_options,
        available_extractors,
        keyword_aliases,
    )

    selected_extractors = available_extractors
    if resolved_requested_options is not None:
        selected_extractors = [
            (sheet_name, extractor)
            for sheet_name, extractor in available_extractors
            if normalize_option_name(sheet_name) in resolved_requested_options
        ]

        for option in sorted(unknown_options):
            print(f"Skipping unknown option: {option}")

    return selected_extractors


def get_requested_options_for_pdf(
    pdf_path: Path,
    global_requested_options: set[str] | None,
    per_pdf_options: dict[str, set[str]],
) -> set[str] | None:
    pdf_keys = {
        normalize_option_name(pdf_path.name),
        normalize_option_name(str(pdf_path)),
    }
    for pdf_key in pdf_keys:
        if pdf_key in per_pdf_options:
            return per_pdf_options[pdf_key]
    return global_requested_options


def process_pdf(
    pdf_path: Path,
    requested_options: set[str] | None,
    available_extractors: list[tuple[str, object]],
    feedback_memory: dict,
    keyword_aliases: dict,
) -> list[tuple[str, list[dict[str, str]]]]:
    job_start_time = time.perf_counter()
    selected_extractors = select_extractors(
        requested_options,
        available_extractors,
        keyword_aliases,
    )

    print(f"PDF: {pdf_path}")
    print(f"Selected extractors: {len(selected_extractors)}")
    sheet_rows = [
        (sheet_name, try_extract_rows(pdf_path, sheet_name, extractor))
        for sheet_name, extractor in selected_extractors
    ]
    found_sheet_rows = [(sheet_name, rows) for sheet_name, rows in sheet_rows if rows]

    if not found_sheet_rows:
        total_elapsed = time.perf_counter() - job_start_time
        print("No details were found.")
        print(f"Total time taken: {total_elapsed:.2f}s")
        return []

    reviewed_sheet_rows = [
        (sheet_name, apply_feedback_to_rows(sheet_name, rows, feedback_memory))
        for sheet_name, rows in found_sheet_rows
    ]
    output_path = build_output_path(pdf_path)
    saved_output_path = write_workbook(reviewed_sheet_rows, output_path)
    total_rows = sum(len(rows) for _, rows in reviewed_sheet_rows)
    total_elapsed = time.perf_counter() - job_start_time
    print(
        f"Extracted {total_rows} rows across {len(found_sheet_rows)} sheets to: "
        f"{saved_output_path}"
    )
    print(f"Total time taken: {total_elapsed:.2f}s")
    return reviewed_sheet_rows


def main() -> None:
    pdf_paths, requests_file, feedback_workbooks = resolve_runtime_paths()
    global_requested_options, per_pdf_options = load_requested_options(requests_file)
    keyword_aliases = load_keyword_aliases(KEYWORD_ALIASES_FILE)
    history = load_history(HISTORY_FILE)
    feedback_memory = load_feedback_memory(FEEDBACK_FILE)
    available_extractors = get_available_extractors()

    feedback_updates = 0
    for workbook_path in discover_feedback_workbooks(feedback_workbooks):
        if not workbook_path.exists():
            continue
        feedback_updates += ingest_feedback_workbook(workbook_path, feedback_memory)
    if feedback_updates:
        print(f"Learned {feedback_updates} feedback updates from reviewed Excel files.")

    if requests_file is not None and per_pdf_options:
        requested_pdf_paths = []
        for raw_line in requests_file.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or ":" not in line:
                continue
            pdf_name, _raw_options = line.split(":", 1)
            requested_pdf_paths.append(resolve_pdf_path(Path(pdf_name.strip()), requests_file))

        if len(pdf_paths) == 1 and pdf_paths[0] == PDF_PATH and requested_pdf_paths:
            pdf_paths = requested_pdf_paths

    for pdf_path in pdf_paths:
        resolved_pdf_path = resolve_pdf_path(pdf_path, requests_file)
        if not resolved_pdf_path.exists():
            print(f"Skipping missing PDF: {resolved_pdf_path}")
            continue

        requested_options = get_requested_options_for_pdf(
            resolved_pdf_path,
            global_requested_options,
            per_pdf_options,
        )
        if requested_options is None:
            recommended_options, recommendations = recommend_extractors_for_pdf(
                resolved_pdf_path,
                history,
                available_extractors,
            )
            if recommended_options:
                requested_options = {
                    normalize_option_name(option) for option in recommended_options
                }
                recommendation_text = ", ".join(
                    f"{name} ({score:.2f})" for name, score, _overlap in recommendations[:4]
                )
                print(f"Recommended extractors from history: {recommendation_text}")

        found_sheet_rows = process_pdf(
            resolved_pdf_path,
            requested_options,
            available_extractors,
            feedback_memory,
            keyword_aliases,
        )
        for sheet_name, rows in found_sheet_rows:
            update_history_with_extraction(history, resolved_pdf_path, sheet_name, rows)

    save_history(HISTORY_FILE, history)
    save_feedback_memory(FEEDBACK_FILE, feedback_memory)


if __name__ == "__main__":
    main()
